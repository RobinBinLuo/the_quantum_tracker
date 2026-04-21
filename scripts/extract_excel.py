from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
SOURCE_XLSX = ROOT.parent / "量子公司调研 （美股）-20260309.xlsx"
OUTPUT_JSON = ROOT / "data" / "companies.json"
COMPANY_DATA_DIR = ROOT / "data" / "companies"
MARKET_METRICS_JSON = ROOT / "data" / "market_metrics.json"

ROUTE_DEFINITIONS = {
    "superconducting": {
        "slug": "superconducting",
        "label": "超导 Superconducting",
        "shortLabel": "超导",
        "summary": "目前最成熟、产业化推进最快的通用量子计算路线之一，依赖低温超导量子比特、微波控制和大规模工程集成。",
        "advantages": [
            "产业链与工程体系较成熟，芯片制造和控制技术积累深。",
            "门操作速度快，适合推进更复杂的量子电路实验。",
            "容易与云平台、模块化系统和软件生态协同商业化。",
        ],
        "drawbacks": [
            "需要极低温环境，制冷、封装和连线复杂度高。",
            "随着量子比特数增加，控制和误差管理难度迅速上升。",
            "真正实现容错量子计算仍需要大量工程突破。",
        ],
        "maturity": "高",
        "bestFor": "通用门模型、云访问、企业试点",
        "mainChallenge": "扩展性与容错",
        "tags": ["低温超导", "Gate-based", "Cloud", "容错路线"],
    },
    "trapped-ion": {
        "slug": "trapped-ion",
        "label": "离子阱 Trapped-Ion",
        "shortLabel": "离子阱",
        "summary": "以单原子离子作为量子比特，拥有高保真和全连接优势，是高质量门操作路线的重要代表。",
        "advantages": [
            "量子比特一致性强、门保真度高。",
            "天然适合高质量量子逻辑门和较长相干时间。",
            "在算法验证和早期商用场景中表现稳定。",
        ],
        "drawbacks": [
            "门速度相对较慢，规模化控制系统复杂。",
            "工程扩展到更大规模时面临架构和集成挑战。",
            "制造和部署成本仍然较高。",
        ],
        "maturity": "中高",
        "bestFor": "高保真门模型、科研与企业试点",
        "mainChallenge": "门速度与规模化控制",
        "tags": ["高保真", "全连接", "离子阱", "Gate-based"],
    },
    "neutral-atom": {
        "slug": "neutral-atom",
        "label": "中性原子 Neutral Atom",
        "shortLabel": "中性原子",
        "summary": "依靠激光操控中性原子阵列，兼具较好扩展性和灵活拓扑，是近年上升很快的量子硬件方向。",
        "advantages": [
            "原子阵列扩展潜力较强，适合大规模布局。",
            "拓扑结构灵活，可针对特定问题设计排布。",
            "在模拟与优化类问题上具有研究吸引力。",
        ],
        "drawbacks": [
            "控制系统复杂，对激光与光学稳定性要求高。",
            "商业化成熟度仍低于超导和离子阱。",
            "长期容错能力和大规模系统集成仍待验证。",
        ],
        "maturity": "中",
        "bestFor": "量子模拟、优化与前沿研究",
        "mainChallenge": "系统稳定性与商业成熟度",
        "tags": ["原子阵列", "激光控制", "量子模拟", "扩展潜力"],
    },
    "photonic": {
        "slug": "photonic",
        "label": "光子 Photonic",
        "shortLabel": "光子",
        "summary": "使用光子作为信息载体，强调室温、通信友好与网络兼容性，是量子计算和量子通信的重要交叉路线。",
        "advantages": [
            "可在室温或相对温和环境下工作，系统部署想象空间大。",
            "天然适配通信、传感和网络化场景。",
            "在芯片级光学器件和集成光子方向上有潜在成本优势。",
        ],
        "drawbacks": [
            "高质量单光子源、探测器和纠缠操控仍具挑战。",
            "通用容错计算距离成熟商业化仍较远。",
            "生态和开发者工具普遍不如超导与离子阱成熟。",
        ],
        "maturity": "中低",
        "bestFor": "量子通信、光学器件、早期平台化",
        "mainChallenge": "器件一致性与通用计算能力",
        "tags": ["室温", "光量子", "通信友好", "集成光子"],
    },
    "quantum-annealing": {
        "slug": "quantum-annealing",
        "label": "量子退火 Quantum Annealing",
        "shortLabel": "量子退火",
        "summary": "聚焦特定优化问题求解，以专用系统先行落地商业场景，和通用门模型路线并不完全相同。",
        "advantages": [
            "更容易在优化类应用中形成较早商业落地。",
            "整机交付、云平台与行业方案结合度较高。",
            "对特定问题结构能展示实际业务价值。",
        ],
        "drawbacks": [
            "不等同于通用量子计算，适用范围更窄。",
            "结果优势往往依赖问题结构，泛化能力有限。",
            "长期技术叙事容易受到门模型路线对比压力。",
        ],
        "maturity": "中",
        "bestFor": "组合优化、专用系统销售",
        "mainChallenge": "泛化能力与长期路线认知",
        "tags": ["优化", "退火机", "专用系统", "商业落地"],
    },
    "topological": {
        "slug": "topological",
        "label": "拓扑 Topological",
        "shortLabel": "拓扑",
        "summary": "希望通过拓扑保护提高量子比特稳定性，是高潜力但高风险的长期路线，仍处于较早阶段。",
        "advantages": [
            "理论上有望降低纠错成本并提升稳定性。",
            "一旦突破，可能对容错量子计算形成跨代优势。",
            "容易与软件平台和云生态叙事结合。",
        ],
        "drawbacks": [
            "实验验证与工程可行性仍存在不确定性。",
            "离大规模商用系统还有较长距离。",
            "投资与研发周期长，成果兑现节奏慢。",
        ],
        "maturity": "早期",
        "bestFor": "前沿研究、长期容错探索",
        "mainChallenge": "实验验证与工程兑现",
        "tags": ["Majorana", "长期路线", "前沿研究", "容错潜力"],
    },
    "quantum-network-security": {
        "slug": "quantum-network-security",
        "label": "量子网络与安全 Quantum Network / Security",
        "shortLabel": "量子安全",
        "summary": "围绕量子安全通信、后量子密码与网络基础设施展开，更偏应用层和安全层，而非通用量子硬件本身。",
        "advantages": [
            "更接近现实企业需求，容易切入网络安全和通信场景。",
            "商业化路径相对直接，不必等待通用量子计算成熟。",
            "能与政府、国防和关键基础设施需求强绑定。",
        ],
        "drawbacks": [
            "与纯量子计算主线相比，市场叙事容易分散。",
            "技术护城河和长期标准化竞争仍需观察。",
            "部分公司更像安全科技公司而非纯量子公司。",
        ],
        "maturity": "中",
        "bestFor": "网络安全、关键基础设施、通信",
        "mainChallenge": "标准化与市场认知定位",
        "tags": ["PQC", "安全通信", "网络基础设施", "应用层"],
    },
    "enabling-infrastructure": {
        "slug": "enabling-infrastructure",
        "label": "基础设施与代工 Enabling Infrastructure",
        "shortLabel": "基础设施",
        "summary": "提供晶圆制造、封装、器件和生态支撑，虽然不是直接卖量子算力，但属于产业链不可缺的一环。",
        "advantages": [
            "更接近半导体和制造能力，商业模式相对清晰。",
            "可服务多条量子技术路线，受益于整个行业扩张。",
            "不必押注单一算法或硬件范式。",
        ],
        "drawbacks": [
            "估值弹性有时不如纯量子计算公司。",
            "行业景气度和客户集中度可能带来波动。",
            "量子业务在财报中的独立可见性往往有限。",
        ],
        "maturity": "中高",
        "bestFor": "制造、封装、产业链协同",
        "mainChallenge": "量子业务纯度与客户拓展",
        "tags": ["Foundry", "制造", "封装", "产业链"],
    },
}

COMPANY_ROUTE_MAP = {
    "Google": "superconducting",
    "IBM": "superconducting",
    "Microsoft": "topological",
    "IonQ": "trapped-ion",
    "Quantum Computing (QUBT)": "photonic",
    "D-Wave Quantum (QBTS)": "quantum-annealing",
    "Rigetti Computing (RGTI)": "superconducting",
    "Arqit Quantum (ARQQ)": "quantum-network-security",
    "SEALSQ Corp (LAES)": "quantum-network-security",
    "英特尔 (INTC)": "superconducting",
    "霍尼韦尔 (HON)/Quantinuum": "trapped-ion",
    "Infleqtion/\n Churchill Capital Corp X (NASDAQ: CCCX)": "neutral-atom",
    "SkyWater Technology, Inc. (SKYT)": "enabling-infrastructure",
}

COMPANY_ADDITIONAL_ROUTE_MAP = {
    "IonQ": ["quantum-network-security"],
}

COMPANY_DISPLAY_NAME_MAP = {
    "霍尼韦尔 (HON)/Quantinuum": "Quantinuum",
    "Infleqtion/\n Churchill Capital Corp X (NASDAQ: CCCX)": "Infleqtion (INFQ)",
    "英特尔 (INTC)": "Intel (INTC)",
}

COMPANY_ZH_SUMMARIES = {
    "Google": "Google 的量子业务集中在 Google Quantum AI，核心目标是沿超导路线推进通用容错量子计算。它目前更像 Alphabet 内部的长期科研与平台能力：硬件、纠错、Cirq 工具链和云端实验访问都在推进，但量子收入尚未单独商业化披露。",
    "IBM": "IBM Quantum 是最早把量子计算做成企业级平台的公司之一，围绕超导处理器、Qiskit 和 IBM Quantum Network 建立完整生态。它的优势在于软硬件一体化、云端访问和企业合作网络，但量子业务仍嵌入 IBM 整体软件与基础设施战略中。",
    "Microsoft": "Microsoft 通过 Azure Quantum 布局量子软件、云平台和合作伙伴硬件，同时押注拓扑量子比特作为长期容错路线。短期看，它更像 Azure 科学计算与 AI 基础设施的延伸；长期看，量子被定位为云计算的下一代加速能力。",
    "IonQ": "IonQ 是美股中最典型的离子阱量子计算纯标的之一，业务覆盖云端访问、系统销售、政府项目和企业合作。公司在收入增速和现金储备上已经进入更高能见度阶段，同时也把量子网络作为重要延伸方向。",
    "Quantum Computing (QUBT)": "Quantum Computing Inc. 以光子技术、薄膜铌酸锂芯片、量子安全通信和优化设备为主线。公司仍处于早期商业化阶段，收入基数很小，但通过 foundry、并购和大额融资试图把光子器件能力转化为可销售产品。",
    "D-Wave Quantum (QBTS)": "D-Wave 是量子退火路线的代表公司，重点不是通用门模型，而是面向优化问题的专用系统、云服务和行业方案。它已经能产生真实商业收入和订单，但市场仍需要判断退火路线的适用边界和长期规模化空间。",
    "Rigetti Computing (RGTI)": "Rigetti 是超导量子计算纯标的，长期围绕 QPU、云访问和整机部署推进。公司技术路线清晰，但收入仍主要来自研发合同和早期系统部署，投资判断高度依赖硬件里程碑兑现与现金使用效率。",
    "Arqit Quantum (ARQQ)": "Arqit 并不是通用量子计算硬件公司，而是量子安全加密和后量子密码迁移标的。它的核心看点在于国防、电信和企业安全客户是否会加速采用量子安全方案，而不是量子处理器性能。",
    "SEALSQ Corp (LAES)": "SEALSQ 以安全芯片、身份认证和后量子密码产品为核心，量子叙事主要体现在量子抗性芯片和主权安全基础设施。它更接近网络安全与半导体安全公司，收入兑现取决于量子安全产品落地节奏。",
    "Intel (INTC)": "Intel 的量子布局依托半导体制造经验和硅自旋量子比特研究，本质上是大型芯片公司的长期技术储备。量子业务目前对财务贡献有限，但如果硅基量子器件路线跑通，Intel 的制造能力可能成为关键优势。",
    "Quantinuum": "Quantinuum 是离子阱路线的头部公司，由 Honeywell Quantum Solutions 与 Cambridge Quantum 组合而来，覆盖硬件、软件和量子安全。它目前仍非独立上市公司，但融资、Helios 系统和潜在 IPO 让其成为最值得跟踪的私有量子公司之一。",
    "Infleqtion (INFQ)": "Infleqtion 以中性原子技术为核心，业务横跨量子计算、精密计时、传感和政府/航天任务。上市后它的关键看点是能否把中性原子平台从科研系统推进到可重复交付的收入模型。",
    "SkyWater Technology, Inc. (SKYT)": "SkyWater 不是纯量子计算公司，而是量子产业链中的制造与代工基础设施标的。它的价值在于为量子芯片、控制器和特殊半导体器件提供制造支持，因此更适合作为量子供应链公司来观察。",
}

ROUTE_I18N_EN = {
    "superconducting": {
        "label": "Superconducting",
        "shortLabel": "Superconducting",
        "summary": "One of the most mature and fastest-commercializing paths for general-purpose quantum computing, built around cryogenic superconducting qubits, microwave control and large-scale systems engineering.",
        "advantages": [
            "A relatively mature engineering stack with deep experience in chip fabrication and control electronics.",
            "Fast gate operations, making it suitable for increasingly complex quantum circuit experiments.",
            "Naturally pairs with cloud access, modular systems and software ecosystems.",
        ],
        "drawbacks": [
            "Requires ultra-low-temperature operation, making cooling, packaging and wiring difficult.",
            "Control and error-management complexity rises quickly as qubit counts scale.",
            "Fault-tolerant quantum computing still requires major engineering breakthroughs.",
        ],
        "maturity": "High",
        "bestFor": "Universal gate-model systems, cloud access, enterprise pilots",
        "mainChallenge": "Scalability and fault tolerance",
        "tags": ["Cryogenic superconducting", "Gate-based", "Cloud", "Fault tolerance"],
    },
    "trapped-ion": {
        "label": "Trapped-Ion",
        "shortLabel": "Trapped-Ion",
        "summary": "Uses individual trapped ions as qubits. The route is known for high fidelity and strong connectivity, making it an important high-quality gate-model architecture.",
        "advantages": [
            "Highly uniform qubits and strong gate fidelity.",
            "Well suited for high-quality logic gates and long coherence times.",
            "Stable performance in algorithm validation and early commercial pilots.",
        ],
        "drawbacks": [
            "Gate speeds are generally slower, and large-scale control systems are complex.",
            "Scaling to much larger machines creates architectural and integration challenges.",
            "Manufacturing and deployment costs remain high.",
        ],
        "maturity": "Medium-high",
        "bestFor": "High-fidelity gate-model systems, research and enterprise pilots",
        "mainChallenge": "Gate speed and scalable control",
        "tags": ["High fidelity", "All-to-all connectivity", "Trapped ions", "Gate-based"],
    },
    "neutral-atom": {
        "label": "Neutral Atom",
        "shortLabel": "Neutral Atom",
        "summary": "Relies on laser-controlled neutral-atom arrays. It combines strong scaling potential with flexible topology and has become a fast-rising hardware direction.",
        "advantages": [
            "Atom arrays have strong scaling potential.",
            "Flexible topology can be adapted to specific problem structures.",
            "Attractive for quantum simulation and optimization research.",
        ],
        "drawbacks": [
            "The control stack is complex and depends on stable laser and optical systems.",
            "Commercial maturity is still behind superconducting and trapped-ion platforms.",
            "Long-term fault tolerance and large-system integration remain to be proven.",
        ],
        "maturity": "Medium",
        "bestFor": "Quantum simulation, optimization and frontier research",
        "mainChallenge": "System stability and commercial maturity",
        "tags": ["Atom arrays", "Laser control", "Quantum simulation", "Scaling potential"],
    },
    "photonic": {
        "label": "Photonic",
        "shortLabel": "Photonic",
        "summary": "Uses photons as information carriers, emphasizing room-temperature potential, communication compatibility and network integration.",
        "advantages": [
            "Can operate at room temperature or less extreme conditions, opening deployment flexibility.",
            "Naturally aligned with communications, sensing and networked systems.",
            "Integrated photonics may offer long-term device and cost advantages.",
        ],
        "drawbacks": [
            "High-quality single-photon sources, detectors and entanglement control remain challenging.",
            "Universal fault-tolerant computing is still far from mature commercialization.",
            "Developer tooling and ecosystem maturity generally lag superconducting and trapped-ion routes.",
        ],
        "maturity": "Medium-low",
        "bestFor": "Quantum communications, optical devices, early platform plays",
        "mainChallenge": "Device consistency and universal-computing capability",
        "tags": ["Room temperature", "Photonic qubits", "Communications", "Integrated photonics"],
    },
    "quantum-annealing": {
        "label": "Quantum Annealing",
        "shortLabel": "Quantum Annealing",
        "summary": "Focuses on solving specific optimization problems with purpose-built systems. It is distinct from universal gate-model quantum computing.",
        "advantages": [
            "Can reach commercial optimization use cases earlier than universal systems.",
            "Combines full-system delivery, cloud access and industry solutions.",
            "Can show business value when the problem structure fits the machine.",
        ],
        "drawbacks": [
            "It is not universal quantum computing and has a narrower application scope.",
            "Advantage depends heavily on the problem structure.",
            "The long-term narrative faces comparison pressure from gate-model systems.",
        ],
        "maturity": "Medium",
        "bestFor": "Combinatorial optimization and specialized systems",
        "mainChallenge": "Generalizability and market perception",
        "tags": ["Optimization", "Annealing systems", "Specialized hardware", "Commercial deployment"],
    },
    "topological": {
        "label": "Topological",
        "shortLabel": "Topological",
        "summary": "Aims to use topological protection to improve qubit stability. It is a high-potential but high-risk long-term route.",
        "advantages": [
            "Could theoretically reduce error-correction overhead and improve stability.",
            "A successful breakthrough could create a generational advantage in fault tolerance.",
            "Fits naturally with software platforms and cloud ecosystem narratives.",
        ],
        "drawbacks": [
            "Experimental validation and engineering feasibility remain uncertain.",
            "Large-scale commercial systems are still far away.",
            "R&D cycles are long and commercialization timing is difficult to predict.",
        ],
        "maturity": "Early",
        "bestFor": "Frontier research and long-term fault-tolerance exploration",
        "mainChallenge": "Experimental proof and engineering delivery",
        "tags": ["Majorana", "Long-term route", "Frontier research", "Fault-tolerance potential"],
    },
    "quantum-network-security": {
        "label": "Quantum Network / Security",
        "shortLabel": "Quantum Security",
        "summary": "Covers quantum-secure communications, post-quantum cryptography and network infrastructure. It is more application- and security-layer oriented than general-purpose quantum hardware.",
        "advantages": [
            "Closer to current enterprise needs in cybersecurity and communications.",
            "Commercialization can proceed without waiting for universal quantum computers.",
            "Connects strongly with government, defense and critical-infrastructure demand.",
        ],
        "drawbacks": [
            "The market narrative can be more diffuse than pure quantum computing.",
            "Technical moats and standardization dynamics still need to be watched.",
            "Some companies are closer to security technology firms than pure quantum companies.",
        ],
        "maturity": "Medium",
        "bestFor": "Cybersecurity, critical infrastructure and communications",
        "mainChallenge": "Standardization and market positioning",
        "tags": ["PQC", "Secure communications", "Network infrastructure", "Application layer"],
    },
    "enabling-infrastructure": {
        "label": "Enabling Infrastructure",
        "shortLabel": "Infrastructure",
        "summary": "Provides wafer manufacturing, packaging, devices and ecosystem support. These firms may not sell quantum compute directly, but they are critical parts of the supply chain.",
        "advantages": [
            "Closer to semiconductor manufacturing, with a clearer business model.",
            "Can serve multiple quantum technology routes and benefit from industry expansion.",
            "Does not need to bet on a single algorithm or hardware paradigm.",
        ],
        "drawbacks": [
            "Valuation upside can be lower than pure-play quantum-computing firms.",
            "Cyclicality and customer concentration can create volatility.",
            "Quantum revenue is often not separately visible in financial reporting.",
        ],
        "maturity": "Medium-high",
        "bestFor": "Manufacturing, packaging and supply-chain collaboration",
        "mainChallenge": "Quantum-business purity and customer expansion",
        "tags": ["Foundry", "Manufacturing", "Packaging", "Supply chain"],
    },
}

COMPANY_EN_CONTENT = {
    "Google": {
        "summary": "Google's quantum effort is centered on Google Quantum AI Lab and its roadmap toward universal fault-tolerant quantum computing. The team develops superconducting processors, error-correction methods and software tools, but the activity is still mainly a research and platform effort rather than a mass-market commercial product.",
        "business": "Current business and product focus\n1. Hardware design and fabrication of superconducting quantum processors, with the goal of building a universal fault-tolerant quantum computer. The latest milestone highlighted in the source material is the 105-qubit Willow chip.\n2. Quantum error correction and fault-tolerant architecture, including surface-code decoding, compilation, scheduling, control circuitry, readout, topology, cooling and packaging.\n3. Quantum software and ecosystem work, including Cirq, quantum algorithms, simulation, optimization, machine learning research and limited cloud access to processors such as Sycamore and Willow.",
        "financial": "Market capitalization: about $3.09T\nP/E (TTM): 27.21",
        "latestQuarter": "Q4 revenue was about $96.5B, up roughly 12% year over year. Full-year revenue was about $348B, up roughly 14%, driven mainly by Google Cloud, AI infrastructure and advertising. Quantum is not reported as a separate revenue segment and was not discussed in the earnings call.",
        "ceoMedia": "Google Quantum AI published a Nature experiment showing quantum processors as platforms for exploring new states of matter. The team also reported Willow T1 times around 100 microseconds, roughly five times the prior generation. Management has described Willow as an important step toward practical quantum computing for areas such as drug discovery, fusion energy and next-generation batteries.",
    },
    "IBM": {
        "summary": "IBM Quantum is one of the most established enterprise quantum platforms, combining superconducting hardware, Qiskit software and cloud access. Its strategy is to build a full-stack ecosystem for research, enterprise pilots and long-term fault-tolerant systems.",
        "business": "Business and product focus\n1. Superconducting quantum processors and modular quantum systems, including cloud-accessible processors and enterprise-facing system deployments.\n2. Qiskit and the broader software stack for circuit design, compilation, error mitigation and workflow integration.\n3. IBM Quantum Network, partnerships and industry programs that connect universities, labs and enterprise customers to IBM's quantum roadmap.",
        "financial": "Quantum is not disclosed as a standalone revenue line. IBM's reported financials are driven by software, consulting and infrastructure, with quantum positioned as a strategic technology platform.",
        "latestQuarter": "The latest quarter material focuses on IBM's broader business rather than standalone quantum revenue. Quantum remains embedded in IBM's long-term platform, cloud and research strategy.",
        "ceoMedia": "IBM continues to communicate its quantum roadmap through IBM Quantum announcements, Qiskit updates, research publications and ecosystem partnerships. Management positions quantum as part of IBM's long-term hybrid-cloud and high-performance computing strategy.",
    },
    "Microsoft": {
        "summary": "Microsoft's quantum work is organized around Azure Quantum and a long-term topological-qubit roadmap. The company emphasizes resource estimation, quantum software, chemistry/materials workflows and a future path toward fault-tolerant quantum computing.",
        "business": "Business and product focus\n1. Azure Quantum as a cloud platform that connects users to quantum hardware partners, simulation tools and workflow services.\n2. Topological qubit research, including Majorana-based approaches intended to reduce error-correction overhead over the long run.\n3. Quantum software, resource estimation and quantum chemistry/materials tooling through Azure Quantum Elements.",
        "financial": "Quantum is not separately disclosed in Microsoft's financial reporting. The activity sits within Microsoft's cloud, AI and research ecosystem.",
        "latestQuarter": "The latest quarter material mainly reflects Microsoft's broader cloud and AI business. Quantum remains a strategic R&D and platform effort rather than a reported standalone segment.",
        "ceoMedia": "Microsoft has highlighted Azure Quantum, resource estimation and topological-qubit work as part of its broader scientific computing strategy. Public messaging emphasizes long-term fault tolerance and integration with cloud services.",
    },
    "IonQ": {
        "summary": "IonQ is a public pure-play quantum-computing company built around trapped-ion hardware. It sells cloud access, enterprise projects and system deployments while also expanding into quantum networking.",
        "business": "Business and product focus\n1. Trapped-ion quantum computers with high-fidelity gates and cloud access through major cloud platforms.\n2. Enterprise and government projects, including algorithm pilots, hybrid workflows and direct system sales.\n3. Quantum networking initiatives, including collaborations with research institutions and programs aimed at connecting quantum systems.",
        "financial": "IonQ is still in a high-growth, investment-heavy stage. Revenue is growing from a small base, while operating losses remain meaningful as the company invests in R&D, commercialization and system deployment.",
        "latestQuarter": "Recent reporting highlights accelerating quantum sales and annual revenue crossing the $100M level. The company continues to emphasize bookings, system contracts and government/enterprise demand as key progress indicators.",
        "ceoMedia": "IonQ frequently communicates progress through system-roadmap updates, government program awards, university collaborations and customer announcements. Management frames trapped-ion performance and networking capability as core differentiators.",
        "notes": "IonQ has also been added to the Quantum Network / Security route because its networking collaborations are increasingly relevant to that category.",
    },
    "Quantum Computing (QUBT)": {
        "summary": "Quantum Computing Inc. positions itself around photonic quantum technologies, quantum-secure communications and related hardware/software products. Its business mix includes early-stage product commercialization, acquisitions and capital raising.",
        "business": "Business and product focus\n1. Photonic and optical quantum technologies, including quantum-secure communications and related components.\n2. Product and platform development around entropy, encryption, optimization and optical quantum systems.\n3. M&A-driven expansion, including acquired capabilities intended to broaden the company's hardware and security portfolio.",
        "financial": "The company is early-stage, with revenue still small relative to operating costs. Recent updates point to rising revenue but also continued expense growth.",
        "latestQuarter": "Recent company news highlights revenue growth, operating-cost increases, leadership changes and acquisition activity. The investment case remains tied to whether photonic and security products can convert into repeatable commercial revenue.",
        "ceoMedia": "Public communication has focused on product launches, acquisitions, CEO changes and demonstrations of quantum-secure communications.",
    },
    "D-Wave Quantum (QBTS)": {
        "summary": "D-Wave is a quantum company focused on quantum annealing and optimization, with a long operating history compared with many peers. It sells cloud access, QCaaS and enterprise optimization solutions rather than pursuing only a universal gate-model roadmap.",
        "business": "Business and product focus\n1. Quantum annealing systems for optimization problems and hybrid quantum-classical workflows.\n2. Leap cloud access and QCaaS contracts for enterprise and government customers.\n3. Industry solutions in logistics, defense planning, scheduling and other optimization-heavy use cases.",
        "financial": "D-Wave reports commercial revenue and bookings, but remains in a growth and investment phase. The source material highlights strong revenue growth and accelerating bookings.",
        "latestQuarter": "Recent reporting points to a 179% revenue jump in 2025, stronger bookings, a major QCaaS contract and expanded headquarters/R&D activity. The central question is whether optimization use cases can scale into durable enterprise demand.",
        "ceoMedia": "D-Wave communications emphasize real-world optimization deployments, customer contracts, defense and enterprise use cases, and the distinction between annealing systems and gate-model quantum computing.",
    },
    "Rigetti Computing (RGTI)": {
        "summary": "Rigetti is a public pure-play superconducting quantum-computing company. It develops quantum processors, cloud access and full-stack systems, while the business remains highly dependent on R&D execution and external funding.",
        "business": "Business and product focus\n1. Superconducting quantum processors and QPUs for cloud and direct system deployments.\n2. Quantum cloud access, software tools and integration with research/enterprise workflows.\n3. International system sales and partnerships, including announced orders and deployments of larger-qubit systems.",
        "financial": "Rigetti remains loss-making and investment-intensive. The source material references a large annual loss while the company continues expanding quantum-system deployments.",
        "latestQuarter": "Recent updates include general availability of a 108-qubit system, international deployments and investment plans. The core execution issue is whether technical progress can translate into durable revenue and improved margins.",
        "ceoMedia": "Rigetti's messaging focuses on superconducting hardware progress, QPU deployments, system orders and its vision for full-stack quantum computing.",
    },
    "Arqit Quantum (ARQQ)": {
        "summary": "Arqit is positioned around quantum-safe encryption and post-quantum security rather than building general-purpose quantum computers. Its commercial story depends on enterprise, telecom and government adoption of quantum-safe security products.",
        "business": "Business and product focus\n1. Quantum-safe encryption software and key-agreement products for telecom, defense and enterprise customers.\n2. Post-quantum cryptography planning and encryption intelligence tools.\n3. Partnerships and pilots with network operators, security providers and government-related customers.",
        "financial": "Revenue remains modest and volatile. Recent updates point to revenue softness but potentially meaningful U.S. defense and telecom opportunities.",
        "latestQuarter": "The latest quarter data in the source file is limited. Recent news emphasizes a registered direct offering, telecom partnerships and post-quantum planning products.",
        "ceoMedia": "Arqit public updates focus on quantum-safe encryption partnerships, telecom tests, defense wins and capital-market activity.",
    },
    "SEALSQ Corp (LAES)": {
        "summary": "SEALSQ focuses on post-quantum security, secure chips and quantum-resilient infrastructure. It is best viewed as a cybersecurity and semiconductor-security company with a quantum-safe positioning.",
        "business": "Business and product focus\n1. Secure semiconductors, identity and post-quantum cryptography products.\n2. Quantum-resilient infrastructure initiatives and investments in quantum-related startups.\n3. Partnerships aimed at adding post-quantum security to drones, FPGAs and connected devices.",
        "financial": "The company has released preliminary 2025 financial metrics, but the quantum contribution is not clearly separated. The business should be evaluated through security-chip demand, partnerships and balance-sheet capacity.",
        "latestQuarter": "Recent updates include expansion of a quantum fund, post-quantum partnerships, investment in EeroQ and preliminary 2025 metrics.",
        "ceoMedia": "SEALSQ messaging emphasizes sovereign quantum infrastructure, post-quantum cryptography, CMOS-compatible quantum architectures and device-level security.",
    },
    "Intel (INTC)": {
        "summary": "Intel's quantum work is tied to semiconductor manufacturing and silicon spin-qubit research. It is not a pure-play quantum company, but its fabrication capabilities could matter if silicon-based quantum devices scale.",
        "business": "Business and product focus\n1. Silicon spin-qubit and quantum-dot devices that leverage Intel's semiconductor manufacturing background.\n2. Research partnerships and test-chip development, including devices used by national labs.\n3. Long-term exploration of scalable quantum hardware through silicon manufacturing processes.",
        "financial": "Quantum is not separately disclosed and is immaterial relative to Intel's core semiconductor business.",
        "latestQuarter": "Recent quantum-specific material highlights deployment of a 12-qubit quantum-dot device with Argonne. Financial reporting remains dominated by Intel's core semiconductor operations.",
        "ceoMedia": "Intel's quantum communication is largely research- and partnership-driven, with emphasis on silicon compatibility and manufacturing scalability.",
    },
    "Quantinuum": {
        "summary": "Quantinuum is a leading trapped-ion quantum company formed from Honeywell Quantum Solutions and Cambridge Quantum. It combines hardware, software and cybersecurity products, while remaining linked to Honeywell ownership and potential IPO plans.",
        "business": "Business and product focus\n1. Trapped-ion quantum computers and protected logical-qubit research.\n2. Quantum software, algorithms and AI-assisted quantum algorithm development.\n3. Cybersecurity and quantum-resilient products, along with international R&D expansion.",
        "financial": "Quantinuum is not separately listed in public markets. Honeywell has indicated that Quantinuum filed confidential paperwork for a possible IPO, so future disclosures may become more detailed.",
        "latestQuarter": "Recent updates include global expansion to Singapore, protected logical-qubit demonstrations, AI-assisted quantum algorithm work and discussion of a possible IPO.",
        "ceoMedia": "Quantinuum public updates focus on the Helios system, logical-qubit progress, international partnerships and commercialization of both hardware and software.",
    },
    "Infleqtion (INFQ)": {
        "summary": "Infleqtion is a neutral-atom quantum company covering computing, sensing and precision timing. After its public-market transition, the company is emphasizing revenue guidance and mission-critical quantum applications.",
        "business": "Business and product focus\n1. Neutral-atom quantum computing systems, including a 100-qubit operational system in the UK.\n2. Quantum precision timing and sensing for mission-critical systems.\n3. Government, aerospace and research partnerships, including NASA-related quantum capabilities.",
        "financial": "The company announced 2026 revenue guidance of $40M and received significant SPAC-related funding, according to recent news. It remains an early-stage public quantum company.",
        "latestQuarter": "Recent updates include NYSE trading, revenue guidance, NASA-related delivery and neutral-atom system deployment in the UK.",
        "ceoMedia": "Infleqtion messaging emphasizes neutral-atom scalability, quantum timing, government/aerospace use cases and operational quantum systems.",
    },
    "SkyWater Technology, Inc. (SKYT)": {
        "summary": "SkyWater is an enabling-infrastructure company rather than a pure quantum-computing vendor. Its relevance comes from semiconductor foundry, fabrication and packaging capabilities that can support quantum devices.",
        "business": "Business and product focus\n1. Semiconductor foundry and advanced manufacturing services.\n2. Collaboration with quantum companies on hybrid quantum-classical chips and superconducting controllers.\n3. Fab and packaging capabilities that support the broader quantum supply chain.",
        "financial": "SkyWater's financials are driven by its broader foundry business. Quantum-related revenue is not separately visible.",
        "latestQuarter": "Recent updates include fourth-quarter and full-year 2025 results, a Fab 25 deal and quantum-related collaborations with Silicon Quantum Computing and QuamCore.",
        "ceoMedia": "Public messaging highlights manufacturing partnerships, supply-chain positioning and specialty foundry capabilities for emerging technologies.",
    },
}

COMPANY_MARKET_DATA = {
    "Google": {
        "ticker": "NASDAQ:GOOGL",
        "newsQuery": "(Alphabet OR Google Quantum AI OR Willow quantum chip)",
        "newsSearchTerm": "Google Quantum AI",
    },
    "IBM": {
        "ticker": "NYSE:IBM",
        "newsQuery": "(IBM Quantum OR IBM quantum computing OR Qiskit)",
        "newsSearchTerm": "IBM Quantum",
    },
    "Microsoft": {
        "ticker": "NASDAQ:MSFT",
        "newsQuery": "(Microsoft quantum OR Azure Quantum OR Majorana quantum)",
        "newsSearchTerm": "Azure Quantum",
    },
    "IonQ": {
        "ticker": "NYSE:IONQ",
        "newsQuery": "(IonQ OR trapped-ion quantum computing)",
        "newsSearchTerm": "IonQ",
    },
    "Quantum Computing (QUBT)": {
        "ticker": "NASDAQ:QUBT",
        "newsQuery": "(Quantum Computing Inc OR QUBT OR photonic quantum computing)",
        "newsSearchTerm": "Quantum Computing Inc",
    },
    "D-Wave Quantum (QBTS)": {
        "ticker": "NYSE:QBTS",
        "newsQuery": "(D-Wave OR D-Wave Quantum OR quantum annealing)",
        "newsSearchTerm": "D-Wave",
    },
    "Rigetti Computing (RGTI)": {
        "ticker": "NASDAQ:RGTI",
        "newsQuery": "(Rigetti OR Rigetti Computing OR superconducting quantum computing)",
        "newsSearchTerm": "Rigetti",
    },
    "Arqit Quantum (ARQQ)": {
        "ticker": "NASDAQ:ARQQ",
        "newsQuery": "(Arqit OR Arqit Quantum OR quantum encryption)",
        "newsSearchTerm": "Arqit",
    },
    "SEALSQ Corp (LAES)": {
        "ticker": "NASDAQ:LAES",
        "newsQuery": "(SEALSQ OR LAES OR post-quantum security)",
        "newsSearchTerm": "SEALSQ",
    },
    "英特尔 (INTC)": {
        "ticker": "NASDAQ:INTC",
        "newsQuery": "(Intel quantum computing OR Tunnel Falls qubits OR Intel quantum)",
        "newsSearchTerm": "Intel quantum",
    },
    "霍尼韦尔 (HON)/Quantinuum": {
        "ticker": "NASDAQ:HON",
        "newsQuery": "(Quantinuum OR Honeywell quantum OR trapped-ion quantum)",
        "newsSearchTerm": "Quantinuum",
    },
    "Infleqtion/\n Churchill Capital Corp X (NASDAQ: CCCX)": {
        "ticker": "NYSE:INFQ",
        "newsQuery": "(Infleqtion OR INFQ OR neutral atom quantum)",
        "newsSearchTerm": "Infleqtion",
    },
    "SkyWater Technology, Inc. (SKYT)": {
        "ticker": "NASDAQ:SKYT",
        "newsQuery": "(SkyWater Technology OR quantum foundry OR SkyWater quantum)",
        "newsSearchTerm": "SkyWater quantum",
    },
}

EARNINGS_REPORTS_EN = {
    "Google": {
        "fy2024_notes": "Alphabet's 2024 annual-call discussion referenced quantum computing as one of Google's next-generation technology frontiers, alongside AI agents, reasoning/deep research and advanced video generation. Quantum remains a strategic research platform rather than a separate commercial reporting segment.",
        "q1_2025": "The Q1 2025 materials did not include a meaningful quantum-computing update. The quarter should be read mainly through Alphabet's broader cloud, AI infrastructure and advertising performance.",
        "q2_2025": "The Q2 2025 earnings release and call did not meaningfully discuss quantum computing. No standalone quantum revenue or operating metric was disclosed.",
        "q3_2025": "The Q3 2025 call mentioned Google's Willow quantum-chip breakthrough. Management highlighted a verifiable algorithmic result that ran about 13,000 times faster than a leading supercomputer, framing the result as a step toward future practical applications.",
        "q4_2025": "Q4 revenue was about $96.5B, up roughly 12% year over year, while full-year revenue was about $348B, up roughly 14%. Growth was driven mainly by Google Cloud, AI infrastructure and advertising. Quantum was not reported as a standalone revenue segment and was not meaningfully discussed on the call.",
    },
    "IBM": {
        "fy2024_notes": "IBM's 2024 annual materials highlighted Guardium Quantum Safe, a product designed to help enterprises identify cryptographic assets, assess quantum-era vulnerabilities and plan encryption upgrades. This reinforces IBM's positioning in both quantum computing and quantum-safe security.",
        "q1_2025": "Q1 2025 financial materials did not disclose standalone quantum revenue. The earnings call referenced deployment of Europe's first IBM Quantum System Two in Spain, using a 156-qubit Heron processor, with installation expected around year-end 2025.",
        "q2_2025": "Q2 2025 financial materials did not disclose standalone quantum revenue. The call noted IBM Quantum System Two deployment with RIKEN in Kobe, Japan, also using a 156-qubit Heron processor.",
        "q3_2025": "Q3 revenue was about $16.5B, up roughly 9% year over year, with software, consulting and infrastructure all contributing. The quarter emphasized IBM's broader software and AI momentum while quantum remained embedded in the long-term platform roadmap.",
        "q4_2025": "Q4 revenue was about $19.7B, and FY2025 revenue was about $67.5B, up roughly 8%. Software and infrastructure were the strongest growth areas, and IBM's GenAI book of business reached about $12.5B. Quantum remained strategically important but was not separated as an independent revenue line.",
    },
    "Microsoft": {
        "fy2024_notes": "Microsoft's annual-call commentary positioned quantum as a future cloud accelerator after AI. Management also referenced progress with Atom Computing and the broader path toward fault-tolerant quantum systems.",
        "q1_2025": "Q1 materials referenced Majorana One and Azure Quantum progress. Microsoft framed quantum as a new frontier for cloud systems, combining its quantum stack with partner hardware and long-term topological-qubit work.",
        "q2_2025": "Q2 materials did not include a direct quantum-computing update. The quarter should be read mainly through Azure, AI infrastructure and Microsoft Cloud performance.",
        "q3_2025": "Q3 revenue was about $70.1B, up about 13%, with Intelligent Cloud revenue of about $26.8B. Management again referenced quantum as a next frontier for cloud systems and linked Majorana-1 to the long-term path toward scalable quantum computing.",
        "q4_2025": "FY2025 Q4 revenue was about $76.4B, up 18%, and full-year revenue was about $281.7B, up 15%. Satya Nadella described quantum as the next major cloud accelerator. Azure Quantum remains the main commercial interface, while Microsoft continues to rely on partner hardware and its long-term Majorana roadmap.",
    },
    "IonQ": {
        "fy2024_notes": "IonQ's 2024 materials highlighted a $54.5M contract with the U.S. Air Force Research Laboratory, 2025 revenue guidance of $75M-$95M and high customer concentration, with two customers contributing about 77% of 2024 revenue.",
        "q1_2025": "Q1 revenue was about $7.6M, with operating costs and expenses of about $83.2M and a net loss of about $32.3M. Reported net loss benefited from a non-cash warrant-liability fair-value gain.",
        "q2_2025": "Q2 revenue was about $20.7M, while operating costs rose sharply to about $181.3M. Net loss was about $177.5M. The quarter showed strong top-line momentum but also heavy investment in hardware, platform expansion and commercialization.",
        "q3_2025": "Q3 revenue was about $39.9M, up roughly 222% year over year. GAAP net loss was about $1.1B, mainly driven by accounting and one-time items rather than operating cash burn. IonQ raised 2025 revenue guidance to about $106M-$110M and highlighted Tempo, quantum networking and hybrid quantum-HPC progress.",
        "q4_2025": "Q4 revenue was about $61.9M, up roughly 429% year over year, and FY2025 revenue reached about $130M. IonQ became the first public quantum-computing company to exceed $100M in annual GAAP revenue. Reported GAAP net income was driven largely by non-cash fair-value adjustments, while adjusted EBITDA remained negative.",
    },
    "Quantum Computing (QUBT)": {
        "fy2024_notes": "QUBT's 2024 materials emphasized its Tempe TFLN photonic integrated-circuit foundry, the Dirac-3 optimization machine and NASA-related validation work. The company positioned photonic manufacturing as the foundation for its commercial quantum and photonic product portfolio.",
        "q1_2025": "Q1 revenue was about $39K, with operating expenses of about $8.3M. Reported net income of about $17M was mainly driven by a non-cash warrant-liability gain. The key business update was commercial start-up of the Tempe TFLN foundry and early product/customer activity.",
        "q2_2025": "Q2 revenue was about $61K, operating expenses were about $10.2M and net loss was about $36.5M, including a large non-cash warrant-derivative remeasurement loss. The quarter showed very early revenue scale but continued investment in photonic manufacturing and device commercialization.",
        "q4_2025": "Q4 revenue was about $198K, up roughly 219% year over year, while FY2025 revenue was about $0.68M. Net loss narrowed meaningfully versus FY2024, but operating expenses expanded due to R&D, manufacturing capability and M&A costs. Cash increased substantially after a $750M private placement.",
    },
    "D-Wave Quantum (QBTS)": {
        "fy2024_notes": "D-Wave ended 2024 with more than $300M of cash and emphasized customer work in optimization use cases such as drug discovery, portfolio optimization and vehicle deployment. The company also advanced Advantage2, its next-generation annealing processor.",
        "q1_2025": "Q1 revenue was about $15.0M, up more than 500% year over year, driven mainly by a first full-system Advantage sale. GAAP gross margin was above 90%, bookings were about $1.6M and adjusted EBITDA remained negative. D-Wave also emphasized a claimed quantum-advantage result in materials simulation.",
        "q2_2025": "Q2 revenue was about $3.1M, up roughly 42% year over year but down sharply from Q1's system-sale-driven level. The company completed a $400M ATM equity raise and ended the quarter with about $819M in cash and equivalents. Advantage2 general availability was the key technology milestone.",
        "q3_2025": "Q3 revenue was about $3.7M, up roughly 100% year over year, and bookings were about $2.4M. GAAP net loss was about $140M, mainly due to non-cash warrant-liability remeasurement. Cash exceeded $836M, giving the company a large runway for R&D and commercialization.",
        "q4_2025": "Q4 revenue was about $2.8M, up roughly 19% year over year, and FY2025 revenue reached about $24.6M, up 179%. GAAP net loss widened for the full year, while gross margin remained high. The main debate is whether annealing and hybrid optimization demand can scale into recurring enterprise revenue.",
    },
    "Rigetti Computing (RGTI)": {
        "fy2024_notes": "Rigetti's 2024 materials framed 2025 as a roadmap year, targeting a system with more than 100 qubits and roughly 2x error-rate improvement. Revenue was still mainly driven by technology-development contracts rather than large-scale quantum-computer sales.",
        "q1_2025": "Q1 revenue was about $1.5M, down roughly 50% year over year. Operating expenses were about $22.1M and operating loss was about $21.6M. Reported net income was driven by non-cash accounting gains. Government contracts and the NQCC system upgrade remained key revenue drivers.",
        "q2_2025": "Q2 revenue was about $1.8M, up modestly from Q1 but down about 40% year over year. Operating loss was about $19.9M and net loss was about $39.7M, including non-cash derivative and earn-out remeasurement losses. Rigetti highlighted its 36-qubit multi-chip quantum processor progress.",
        "q3_2025": "Q3 revenue was about $1.9M and operating loss was about $20.5M. GAAP net loss was about $201M, while non-GAAP net loss was about $10.7M. Cash and marketable securities were about $558.9M at quarter-end, later rising to about $600M including warrant exercise proceeds.",
        "q4_2025": "Q4 revenue was about $1.9M, down roughly 17% year over year, and FY2025 revenue was about $7.1M versus about $10.8M in FY2024. GAAP net loss was about $18.2M and non-GAAP net loss was about $11.3M. Cash and marketable securities reached about $589.8M, supporting the superconducting-hardware roadmap.",
    },
    "Arqit Quantum (ARQQ)": {
        "fy2024_notes": "Arqit's FY2024 revenue was about $293K, down from about $640K in FY2023. The company remained focused on quantum-safe encryption software and early defense/telecom commercialization rather than quantum hardware.",
        "q1_2025": "Q1 materials reported revenue of about $67K and a net loss of about $17.2M. Arqit highlighted embedding its symmetric-key agreement software into a U.S. Department of Defense-funded solution with a major IT supplier.",
        "q2_2025": "Arqit reports on a semiannual cadence, so there was no conventional Q2 quarterly report in the source file. The company should be tracked through half-year revenue, defense wins, telecom partnerships and cash usage.",
    },
    "SEALSQ Corp (LAES)": {
        "fy2024_notes": "SEALSQ expected its new quantum-resistant chip to begin commercial product release around Q4 2025 and broader deployment or revenue contribution in 2026. The company also reported about $6.4M of confirmed 2025 orders and raised roughly $60M through direct offerings to accelerate post-quantum chip deployment.",
        "q1_2025": "H1 2025 materials positioned quantum security as SEALSQ's core strategy. The company delivered QS7001 post-quantum microcontroller samples, planned QVault TPM versions for late 2025/early 2026, guided FY2025 revenue to $17.5M-$20.0M and maintained a large cash base for expansion and acquisitions.",
        "q2_2025": "SEALSQ uses a semiannual reporting cadence, so the Q2 field refers to the same half-year reporting framework rather than a separate U.S.-style quarterly filing.",
        "q3_2025": "Q3 revenue was about $5.1M, and nine-month revenue was about $9.9M, up roughly 41% year over year. The company reiterated FY2025 revenue guidance of $17.5M-$20.0M and reported substantial cash resources. The business remains in a transition phase with high investment in post-quantum security and ASIC capability.",
    },
    "Intel (INTC)": {
        "q1_2025": "Intel's Q1 2025 earnings release and call did not meaningfully mention quantum computing. Quantum remains a long-term research and semiconductor-manufacturing option rather than a financial driver.",
        "q2_2025": "Intel's Q2 2025 earnings release and call also did not meaningfully discuss quantum computing. The quarter should be interpreted mainly through Intel's core semiconductor turnaround and foundry strategy.",
    },
    "Quantinuum": {
        "q1_2025": "Honeywell's Q1 materials mentioned quantum mainly through the expectation that Quantinuum could become a leading quantum-computing IPO candidate over time. No standalone Quantinuum financial detail was disclosed.",
        "q2_2025": "Honeywell's Q2 materials provided limited quantum-specific disclosure. Quantinuum remained part of Honeywell's long-term technology and portfolio strategy rather than a separately reported public company.",
        "q4_2025": "Honeywell's Q4 2025 call provided a substantive Quantinuum update: roughly $840M of financing, an approximately $10B pre-money valuation, the Helios trapped-ion system with 64+ qubits and partnerships with Quanta Computer, NVIDIA, JPMorgan, Amgen and Mitsui. The update supports the view that Honeywell is preparing Quantinuum as a more independent quantum platform.",
    },
    "Infleqtion (INFQ)": {
        "q1_2025": "The Infleqtion and Churchill Capital Corp X presentation disclosed about $29M of trailing-12-month revenue as of June 30, 2025, roughly $50M of awarded/backlog orders and a sales pipeline of more than $300M. These figures frame Infleqtion as an early commercial neutral-atom company moving toward public-market visibility.",
    },
    "SkyWater Technology, Inc. (SKYT)": {
        "q3_2025": "Q3 revenue was about $150.7M, up roughly 60.7% year over year. GAAP gross profit was about $36.2M with a 24.0% gross margin, non-GAAP net income was about $11.5M and GAAP net income was about $144M, largely due to a one-time accounting gain. Q4 revenue guidance was $155M-$165M.",
    },
}


METRIC_VALUE_EN_OVERRIDES = {
    ("Google", "quantum_revenue_2024"): "Not separately disclosed.",
    ("Google", "quantum_spend_2024"): "None disclosed.",
    ("IBM", "quantum_revenue_2024"): "IBM stated in May 2025 that cumulative quantum-business signings/bookings had reached about $1B since Q1 2017. This is cumulative bookings, not annual revenue.",
    ("IBM", "quantum_spend_2024"): "None disclosed.",
    ("IBM", "revenue_2025_guidance"): "At least 5%, based on the Q2 2025 press-release guidance.",
    ("Microsoft", "quantum_revenue_2024"): "Not separately disclosed.",
    ("Microsoft", "quantum_spend_2024"): "None disclosed.",
    ("IonQ", "quantum_revenue_2024"): "Total revenue was $43.1M, nearly doubling from 2023.",
    ("IonQ", "quantum_spend_2024"): "Operating expenses were about $276M. FY2024 net loss was about $331.6M.",
    ("Quantum Computing (QUBT)", "quantum_revenue_2024"): "Total company revenue was about $373K.",
    ("Quantum Computing (QUBT)", "quantum_spend_2024"): "Total operating expenses were about $26.049M.",
    ("D-Wave Quantum (QBTS)", "quantum_revenue_2024"): "FY2024 revenue was $8.827M, including QCaaS $6.745M, professional services $1.938M and other revenue $0.144M.",
    ("D-Wave Quantum (QBTS)", "quantum_spend_2024"): "Cost of revenue was $3.264M; total operating expenses were $82.786M.",
    ("Rigetti Computing (RGTI)", "quantum_revenue_2024"): "Revenue was about $10.79M. Operating loss was $68.5M and FY2024 net loss was about $201.0M.",
    ("Rigetti Computing (RGTI)", "quantum_spend_2024"): "Spending and operations totaled about $74.2M. Rigetti and Quanta Computer also agreed to invest in superconducting quantum computing over five years.",
    ("Arqit Quantum (ARQQ)", "quantum_revenue_2024"): "FY2024 revenue was about $293K. Pretax net loss was about $23.9M.",
    ("SEALSQ Corp (LAES)", "quantum_revenue_2024"): "About $11M total revenue; quantum-related revenue was not disclosed separately.",
    ("SEALSQ Corp (LAES)", "quantum_spend_2024"): "Total 2024 R&D expense was about $5.0M, with part of the spend related to post-quantum products.",
    ("SEALSQ Corp (LAES)", "remaining_spend_2025"): "H1 R&D spend used: $4.7M.",
    ("Intel (INTC)", "quantum_revenue_2024"): "Not separately disclosed.",
    ("Intel (INTC)", "quantum_spend_2024"): "None disclosed.",
    ("Quantinuum", "quantum_revenue_2024"): "None disclosed.",
    ("Quantinuum", "quantum_spend_2024"): "None disclosed.",
    ("SkyWater Technology, Inc. (SKYT)", "quantum_revenue_2024"): "Revenue was about $342.3M, up roughly 19% year over year. GAAP gross profit was about $69.6M, with a gross margin of about 20.3%. GAAP net loss was about $6.8M, while non-GAAP net income was about $2.7M.",
}

METRIC_VALUE_ZH_OVERRIDES = {
    ("Google", "quantum_revenue_2024"): "无",
    ("Google", "quantum_spend_2024"): "无",
    ("IBM", "quantum_spend_2024"): "无",
    ("Microsoft", "quantum_revenue_2024"): "无",
    ("Microsoft", "quantum_spend_2024"): "无",
    ("Intel (INTC)", "quantum_revenue_2024"): "无",
    ("Intel (INTC)", "quantum_spend_2024"): "无",
    ("Quantinuum", "quantum_revenue_2024"): "无",
    ("Quantinuum", "quantum_spend_2024"): "无",
}


HEADER_ALIASES = {
    "行业": "industry",
    "标的": "company",
    "业务和产品的内在价值分析": "business_analysis",
    "2024年报指引+内容": "fy2024_notes",
    "2025 Q1 财报（Earnings Release）+电话会 (Webcast&transcript)": "q1_2025",
    "2025 Q2 财报（Earnings Release）+电话会 (Webcast&transcript)": "q2_2025",
    "2025 Q3 财报（Earnings Release）+电话会 (Webcast&transcript)": "q3_2025",
    "2025 Q4 财报（Earnings Release）+电话会 (Webcast&transcript)": "q4_2025",
    "官媒/CEO信息": "ceo_media_updates",
    "财务信息": "financial_snapshot",
    "24年量子业务收入": "quantum_revenue_2024",
    "24年量子业务支出": "quantum_spend_2024",
    "当前市值": "market_cap",
    "24年收入": "revenue_2024",
    "毛利率 / Non-GAAP": "gross_margin_non_gaap",
    "GAAP": "gross_margin_gaap",
    "25年收入指引": "revenue_2025_guidance",
    "25年收入增速": "revenue_2025_growth",
    "PS": "ps_ratio",
    "EV/Sales": "ev_sales",
    "滚动PE": "ttm_pe",
    "PB": "pb_ratio",
    "预估25年支出": "estimated_spend_2025",
    "25年已完成支出": "completed_spend_2025",
    "本年度剩余\n支出估算": "remaining_spend_2025",
    "净资产": "net_assets",
    "净现金": "net_cash",
    "总融资性债务 / 短期融资性债务": "short_term_financing_debt",
    "长期融资性债务": "long_term_financing_debt",
    "总资产 / 现金及等价物": "cash_and_equivalents",
    "其他流动资产": "other_current_assets",
    "非流动资产": "non_current_assets",
    "评价与备注": "notes",
    "最新财报": "latest_earnings",
    "最新季度报表": "latest_quarterly_report",
    "最新年度报表": "latest_annual_report",
}

TEXT_PRIORITY_FIELDS = [
    "business_analysis",
    "financial_snapshot",
    "notes",
    "ceo_media_updates",
    "q4_2025",
    "q3_2025",
    "q2_2025",
    "q1_2025",
]

NUMERIC_FIELDS = {
    "market_cap",
    "revenue_2024",
    "gross_margin_non_gaap",
    "gross_margin_gaap",
    "ps_ratio",
    "ev_sales",
    "ttm_pe",
    "pb_ratio",
    "net_assets",
}

FINANCIAL_FIELD_LABELS = {
    "quantum_revenue_2024": ("2024 量子业务收入", "2024 Quantum Revenue"),
    "quantum_spend_2024": ("2024 量子业务支出", "2024 Quantum Spend"),
    "market_cap": ("当前市值", "Market Cap"),
    "revenue_2024": ("2024 收入", "2024 Revenue"),
    "gross_margin_non_gaap": ("毛利率 Non-GAAP", "Gross Margin Non-GAAP"),
    "gross_margin_gaap": ("毛利率 GAAP", "Gross Margin GAAP"),
    "revenue_2025_guidance": ("2025 收入指引", "2025 Revenue Guidance"),
    "revenue_2025_growth": ("2025 收入增速", "2025 Revenue Growth"),
    "ps_ratio": ("P/S", "P/S"),
    "ev_sales": ("EV/Sales", "EV/Sales"),
    "ttm_pe": ("滚动 P/E", "TTM P/E"),
    "pb_ratio": ("P/B", "P/B"),
    "estimated_spend_2025": ("预估 2025 支出", "Estimated 2025 Spend"),
    "completed_spend_2025": ("2025 已完成支出", "2025 Completed Spend"),
    "remaining_spend_2025": ("本年度剩余支出估算", "Estimated Remaining Spend"),
    "net_assets": ("净资产", "Net Assets"),
    "net_cash": ("净现金", "Net Cash"),
    "short_term_financing_debt": ("短期融资性债务", "Short-term Financing Debt"),
    "long_term_financing_debt": ("长期融资性债务", "Long-term Financing Debt"),
    "cash_and_equivalents": ("总资产 / 现金及等价物", "Total Assets / Cash and Equivalents"),
    "other_current_assets": ("其他流动资产", "Other Current Assets"),
    "non_current_assets": ("非流动资产", "Non-current Assets"),
}

MARKET_METRIC_LABELS = {
    "marketCap": ("当前市值", "Market Cap"),
    "enterpriseValue": ("企业价值 EV", "Enterprise Value"),
    "ttmRevenue": ("TTM 收入", "TTM Revenue"),
    "grossProfit": ("TTM 毛利", "TTM Gross Profit"),
    "grossMargin": ("TTM 毛利率", "TTM Gross Margin"),
    "ps": ("P/S", "P/S"),
    "evSales": ("EV/Sales", "EV/Sales"),
    "pe": ("滚动 P/E", "TTM P/E"),
    "pb": ("P/B", "P/B"),
    "totalDebt": ("总债务", "Total Debt"),
    "netCash": ("净现金", "Net Cash"),
}


def combine_headers(sheet: Any) -> list[str]:
    headers = []
    for col in range(1, sheet.max_column + 1):
        top = sheet.cell(1, col).value
        sub = sheet.cell(2, col).value
        if top and sub:
            headers.append(f"{top} / {sub}")
        else:
            headers.append(top or sub or f"col_{col}")
    return headers


def clean_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, str):
        normalized = (
            value.replace("\r\n", "\n")
            .replace("\r", "\n")
            .replace("_x000B_", "\n")
            .replace("\x0b", "\n")
            .strip()
        )
        return normalized or None
    return value


def slugify(index: int, company_name: str) -> str:
    ascii_part = re.sub(r"[^a-z0-9]+", "-", company_name.lower()).strip("-")
    ascii_part = re.sub(r"-{2,}", "-", ascii_part)
    return f"c{index:02d}-{ascii_part or 'quantum-company'}"


def company_filename(company_name: str) -> str:
    filename = re.sub(r'[\\/:*?"<>|\n\r]+', "-", company_name).strip()
    filename = re.sub(r"\s+", " ", filename)
    filename = filename.replace(" ", "_")
    filename = filename.strip("._-")
    return f"{filename or 'quantum-company'}.json"


def extract_links(text: str | None) -> list[str]:
    if not text:
        return []
    matches = re.findall(r"https?://[^\s)>\u3002]+", text)
    seen = []
    for match in matches:
        if match not in seen:
            seen.append(match)
    return seen


def find_numeric(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if not isinstance(value, str):
        return None
    text = value.strip().replace(",", "")
    if not text:
        return None
    if re.fullmatch(r"-?\d+(\.\d+)?", text):
        return float(text)
    return None


def load_market_metrics() -> dict[str, Any]:
    if not MARKET_METRICS_JSON.exists():
        return {}
    try:
        payload = json.loads(MARKET_METRICS_JSON.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}
    return payload.get("companies", {})


def normalize_metric_value(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text in {"#VALUE!", "N/A", "n/a", "-"}:
        return None
    return text


def translate_metric_value(value: Any) -> str | None:
    text = normalize_metric_value(value)
    if not text:
        return text
    if not re.search(r"[\u4e00-\u9fff]", text):
        return text

    replacements = [
        ("只找到总研发经费为", "Only total R&D expense was found:"),
        ("这个数字是包含所有研发经费，不只是量子计算", "This figure includes all R&D expense, not only quantum computing"),
        ("约", "about "),
        ("营收", "Revenue"),
        ("收入", "Revenue"),
        ("同比增长", "YoY growth"),
        ("毛利率", "gross margin"),
        ("毛利", "gross profit"),
        ("年度净亏损", "annual net loss"),
        ("净亏损", "net loss"),
        ("净收益", "net income"),
        ("每股", "per share"),
        ("现金及投资合计", "cash and investments"),
        ("现金及等价物", "cash and equivalents"),
        ("现金储备", "cash reserve"),
        ("另：", "also: "),
        ("可售证券", "marketable securities"),
        ("受限现金", "restricted cash"),
        ("7月完成", "completed in July"),
        ("股权融资后备考现金", "pro forma cash after equity financing"),
        ("前九个月", "first nine months"),
        ("上半年", "H1"),
        ("公司指导", "company guidance"),
        ("目标为", "target of"),
        ("预计", "expected"),
        ("增长", "growth"),
        ("包含", "includes"),
        ("所有", "all"),
        ("研发经费", "R&D expense"),
        ("不是", "not"),
        ("量子计算", "quantum computing"),
        ("美元", "USD"),
        ("亿", "00M"),
        ("万", "0K"),
    ]
    translated = text
    for source, target in replacements:
        translated = translated.replace(source, target)
    translated = translated.replace("（", " (").replace("）", ")")
    translated = translated.replace("；", "; ").replace("，", ", ").replace("。", ". ")
    translated = re.sub(r"\s+", " ", translated).strip()
    return translated


def build_financial_metrics(company_name: str, normalized: dict[str, Any], market_entry: dict[str, Any]) -> list[dict[str, Any]]:
    items = []
    market_metrics = market_entry.get("metrics", {}) if market_entry else {}
    market_source = market_entry.get("source", "StockAnalysis") if market_entry else "StockAnalysis"
    market_source_url = market_entry.get("sourceUrl") if market_entry else None

    for key, (label_zh, label_en) in MARKET_METRIC_LABELS.items():
        value = normalize_metric_value(market_metrics.get(key))
        if not value:
            continue
        items.append(
            {
                "key": key,
                "label": label_zh,
                "value": value,
                "source": market_source,
                "sourceUrl": market_source_url,
                "i18n": {"en": {"label": label_en, "value": value}},
            }
        )

    for key, (label_zh, label_en) in FINANCIAL_FIELD_LABELS.items():
        value = normalize_metric_value(normalized.get(key))
        if not value:
            continue
        if key in {"market_cap", "ps_ratio", "ev_sales", "ttm_pe", "pb_ratio", "net_cash"} and any(
            item["key"] in {"marketCap", "ps", "evSales", "pe", "pb", "netCash"} for item in items
        ):
            continue
        items.append(
            {
                "key": key,
                "label": label_zh,
                "value": METRIC_VALUE_ZH_OVERRIDES.get((company_name, key), value),
                "source": "Excel",
                "sourceUrl": None,
                "i18n": {
                    "en": {
                        "label": label_en,
                        "value": METRIC_VALUE_EN_OVERRIDES.get((company_name, key), translate_metric_value(value)),
                    }
                },
            }
        )

    return items


def build_earnings_reports(company_name: str, sections: dict[str, Any]) -> list[dict[str, Any]]:
    report_defs = [
        ("fy2024_notes", "2024 年报", "FY2024 Annual Report"),
        ("q1_2025", "2025 Q1", "2025 Q1"),
        ("q2_2025", "2025 Q2", "2025 Q2"),
        ("q3_2025", "2025 Q3", "2025 Q3"),
        ("q4_2025", "2025 Q4", "2025 Q4"),
    ]
    reports = []
    for key, label_zh, label_en in report_defs:
        content = sections.get(key)
        if not content:
            continue
        reports.append(
            {
                "key": key,
                "label": label_zh,
                "content": content,
                "i18n": {
                    "en": {
                        "label": label_en,
                        "content": EARNINGS_REPORTS_EN.get(company_name, {}).get(
                            key,
                            "No English summary has been prepared for this reporting period yet.",
                        ),
                    }
                },
            }
        )
    return reports


def split_summary_and_fundamentals(text: str | None) -> tuple[str | None, str | None]:
    if not isinstance(text, str):
        return None, text

    normalized = text.strip()
    if not normalized:
        return None, normalized

    split_patterns = [
        r"\n\s*目前主要业务",
        r"\n\s*目前主营业务",
        r"\n\s*主要业务",
        r"\n\s*核心业务",
        r"\n\s*业务模式",
    ]
    split_at = None
    for pattern in split_patterns:
        match = re.search(pattern, normalized)
        if match and (split_at is None or match.start() < split_at):
            split_at = match.start()

    if split_at is None:
        paragraphs = [part.strip() for part in re.split(r"\n\s*\n", normalized) if part.strip()]
        summary = paragraphs[0] if paragraphs else normalized
        return summary, normalized

    summary = normalized[:split_at].strip()
    fundamentals = normalized[split_at:].strip()
    return summary or None, fundamentals or normalized


def build_companies() -> list[dict[str, Any]]:
    workbook = openpyxl.load_workbook(SOURCE_XLSX, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = combine_headers(sheet)
    companies = []
    market_metrics = load_market_metrics()

    for row_index in range(3, sheet.max_row + 1):
        source_company_name = clean_value(sheet.cell(row_index, 2).value)
        if not source_company_name:
            continue
        company_name = COMPANY_DISPLAY_NAME_MAP.get(source_company_name, source_company_name)

        raw = {}
        normalized = {}
        numeric = {}
        link_pool = []

        for col_index, header in enumerate(headers, start=1):
            raw_value = clean_value(sheet.cell(row_index, col_index).value)
            if raw_value is None:
                continue

            alias = HEADER_ALIASES.get(header, header)
            raw[header] = raw_value
            normalized[alias] = raw_value

            numeric_value = find_numeric(raw_value)
            if alias in NUMERIC_FIELDS and numeric_value is not None:
                numeric[alias] = numeric_value

            if isinstance(raw_value, str):
                link_pool.extend(extract_links(raw_value))

        business_summary, business_fundamentals = split_summary_and_fundamentals(normalized.get("business_analysis"))
        if business_fundamentals:
            normalized["business_analysis"] = business_fundamentals

        summary = business_summary
        if not summary:
            for key in TEXT_PRIORITY_FIELDS:
                value = normalized.get(key)
                if isinstance(value, str) and value.strip():
                    summary = value.strip()
                    break

        quarter_key = next(
            (key for key in ["q4_2025", "q3_2025", "q2_2025", "q1_2025"] if normalized.get(key)),
            None,
        )

        route_key = COMPANY_ROUTE_MAP.get(source_company_name, "enabling-infrastructure")
        route = ROUTE_DEFINITIONS[route_key]
        route_en = ROUTE_I18N_EN[route_key]
        market_data = COMPANY_MARKET_DATA.get(source_company_name, {})
        filename = company_filename(company_name)
        company_en = COMPANY_EN_CONTENT.get(company_name, {})
        market_entry = market_metrics.get(company_name, {})
        financial_metrics = build_financial_metrics(company_name, normalized, market_entry)
        earnings_reports = build_earnings_reports(company_name, normalized)
        latest_report = earnings_reports[-1] if earnings_reports else None
        current_market = market_entry.get("metrics", {}) if market_entry else {}
        summary_zh = COMPANY_ZH_SUMMARIES.get(company_name) or polish_summary(summary, company_name, route)

        companies.append(
            {
                "id": slugify(len(companies) + 1, company_name),
                "name": company_name,
                "file": f"./data/companies/{filename}",
                "industry": normalized.get("industry"),
                "summary": summary_zh,
                "i18n": {
                    "en": {
                        "industry": "Quantum",
                        "summary": company_en.get(
                            "summary",
                            f"{company_name} is active in the {route_en['shortLabel']} quantum technology route, with the business still centered on technical progress and commercialization validation.",
                        ),
                    }
                },
                "ticker": market_data.get("ticker"),
                "newsQuery": market_data.get("newsQuery", company_name),
                "newsSearchTerm": market_data.get("newsSearchTerm", company_name),
                "latestQuarterKey": latest_report["key"] if latest_report else quarter_key,
                "latestQuarterLabel": latest_report["label"] if latest_report else {
                    "q4_2025": "2025 Q4",
                    "q3_2025": "2025 Q3",
                    "q2_2025": "2025 Q2",
                    "q1_2025": "2025 Q1",
                }.get(quarter_key),
                "route": {
                    "slug": route["slug"],
                    "label": route["label"],
                    "shortLabel": route["shortLabel"],
                    "summary": route["summary"],
                    "tags": route["tags"],
                    "i18n": {
                        "en": {
                            "label": route_en["label"],
                            "shortLabel": route_en["shortLabel"],
                            "summary": route_en["summary"],
                            "tags": route_en["tags"],
                        }
                    },
                },
                "metrics": {
                    "marketCap": current_market.get("marketCap") or normalized.get("market_cap"),
                    "revenue2024": normalized.get("revenue_2024"),
                    "guidance2025": normalized.get("revenue_2025_guidance"),
                    "growth2025": normalized.get("revenue_2025_growth"),
                    "ps": current_market.get("ps") or normalized.get("ps_ratio"),
                    "evSales": current_market.get("evSales") or normalized.get("ev_sales"),
                    "pe": current_market.get("pe") or normalized.get("ttm_pe"),
                    "pb": current_market.get("pb") or normalized.get("pb_ratio"),
                    "netCash": current_market.get("netCash") or normalized.get("net_cash"),
                },
                "financialMetrics": financial_metrics,
                "marketMetricsSource": {
                    "source": market_entry.get("source"),
                    "sourceUrl": market_entry.get("sourceUrl"),
                    "ticker": market_entry.get("ticker"),
                } if market_entry else None,
                "earningsReports": earnings_reports,
                "numericMetrics": numeric,
                "sections": normalized,
                "sectionsEn": {
                    "business_analysis": company_en.get("business"),
                    "financial_snapshot": company_en.get("financial"),
                    "latest_quarter": company_en.get("latestQuarter"),
                    "ceo_media_updates": company_en.get("ceoMedia"),
                    "notes": company_en.get("notes"),
                },
                "raw": raw,
                "links": sorted(set(link_pool)),
            }
        )

    return companies


def polish_summary(summary: str | None, company_name: str, route: dict[str, Any]) -> str:
    if not summary:
        return f"{company_name} 主要沿着{route['shortLabel']}路线布局量子技术，相关业务仍处于技术推进和商业化验证阶段。"

    cleaned = re.sub(r"\s+", " ", summary).strip()
    cleaned = cleaned.replace("_x000B_", " ")
    if len(cleaned) <= 260:
        return cleaned

    sentences = re.split(r"(?<=[。！？])", cleaned)
    result = "".join(sentence for sentence in sentences[:3]).strip()
    return result or cleaned[:260]


def build_company_index(company: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": company["id"],
        "name": company["name"],
        "file": company["file"],
        "industry": company.get("industry"),
        "summary": company.get("summary"),
        "i18n": company.get("i18n"),
        "ticker": company.get("ticker"),
        "newsSearchTerm": company.get("newsSearchTerm"),
        "latestQuarterLabel": company.get("latestQuarterLabel"),
        "route": company.get("route"),
        "metrics": company.get("metrics"),
        "financialMetrics": company.get("financialMetrics"),
        "marketMetricsSource": company.get("marketMetricsSource"),
        "numericMetrics": company.get("numericMetrics"),
    }


def build_company_detail(company: dict[str, Any]) -> dict[str, Any]:
    sections = company.get("sections", {})
    sections_en = company.get("sectionsEn", {})
    return {
        "id": company["id"],
        "name": company["name"],
        "industry": company.get("industry"),
        "summary": company.get("summary"),
        "i18n": company.get("i18n"),
        "ticker": company.get("ticker"),
        "newsQuery": company.get("newsQuery"),
        "newsSearchTerm": company.get("newsSearchTerm"),
        "latestQuarterKey": company.get("latestQuarterKey"),
        "latestQuarterLabel": company.get("latestQuarterLabel"),
        "route": company.get("route"),
        "metrics": company.get("metrics"),
        "financialMetrics": company.get("financialMetrics"),
        "marketMetricsSource": company.get("marketMetricsSource"),
        "earningsReports": company.get("earningsReports", []),
        "numericMetrics": company.get("numericMetrics"),
        "fundamentalAnalysis": {
            "business": sections.get("business_analysis"),
            "financial": sections.get("financial_snapshot"),
            "latestQuarter": sections.get(company.get("latestQuarterKey")) if company.get("latestQuarterKey") else None,
            "ceoMedia": sections.get("ceo_media_updates"),
            "notes": sections.get("notes"),
            "latestEarnings": sections.get("latest_earnings"),
            "latestQuarterlyReport": sections.get("latest_quarterly_report"),
            "latestAnnualReport": sections.get("latest_annual_report"),
        },
        "fundamentalAnalysisI18n": {
            "en": {
                "business": sections_en.get("business_analysis"),
                "financial": sections_en.get("financial_snapshot"),
                "latestQuarter": sections_en.get("latest_quarter"),
                "ceoMedia": sections_en.get("ceo_media_updates"),
                "notes": sections_en.get("notes"),
            }
        },
        "links": company.get("links", []),
    }


def build_routes(companies: list[dict[str, Any]]) -> list[dict[str, Any]]:
    routes = []
    for route_key, definition in ROUTE_DEFINITIONS.items():
        related = [
            company
            for company in companies
            if company["route"]["slug"] == definition["slug"]
            or definition["slug"] in COMPANY_ADDITIONAL_ROUTE_MAP.get(company["name"], [])
        ]
        if not related:
            continue
        definition_en = ROUTE_I18N_EN[route_key]
        routes.append(
            {
                "slug": definition["slug"],
                "label": definition["label"],
                "shortLabel": definition["shortLabel"],
                "summary": definition["summary"],
                "advantages": definition["advantages"],
                "drawbacks": definition["drawbacks"],
                "maturity": definition["maturity"],
                "bestFor": definition["bestFor"],
                "mainChallenge": definition["mainChallenge"],
                "tags": definition["tags"],
                "i18n": {
                    "en": {
                        "label": definition_en["label"],
                        "shortLabel": definition_en["shortLabel"],
                        "summary": definition_en["summary"],
                        "advantages": definition_en["advantages"],
                        "drawbacks": definition_en["drawbacks"],
                        "maturity": definition_en["maturity"],
                        "bestFor": definition_en["bestFor"],
                        "mainChallenge": definition_en["mainChallenge"],
                        "tags": definition_en["tags"],
                    }
                },
                "companyCount": len(related),
                "companies": [
                    build_company_index(company)
                    for company in related
                ],
            }
        )
    return routes


def build_payload() -> dict[str, Any]:
    companies = build_companies()
    routes = build_routes(companies)
    return {
        "meta": {
            "title": "量子公司调研（美股）",
            "sourceFile": SOURCE_XLSX.name,
            "generatedAt": datetime.now().isoformat(timespec="seconds"),
            "companyCount": len(companies),
            "routeCount": len(routes),
            "notes": "网页中的数值与文字均沿用原始 Excel 口径，部分字段为原表手工整理后的描述。",
        },
        "newsQuery": "(\"quantum computing\" OR \"quantum computer\" OR \"quantum technology\" OR IonQ OR IBM Quantum OR Quantinuum OR Rigetti OR D-Wave)",
        "routes": routes,
        "companies": [build_company_index(company) for company in companies],
    }


def main() -> None:
    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    COMPANY_DATA_DIR.mkdir(parents=True, exist_ok=True)
    for old_file in COMPANY_DATA_DIR.glob("*.json"):
        old_file.unlink()
    payload = build_payload()
    OUTPUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    for company in build_companies():
        output_path = ROOT / company["file"]
        output_path.write_text(json.dumps(build_company_detail(company), ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUTPUT_JSON}")


if __name__ == "__main__":
    main()
